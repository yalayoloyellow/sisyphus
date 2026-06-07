"""
core/logic.py
"""

from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Tuple

from .data import ensure_state

def get_visible_entries(state: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Только заметки и задачи с done=False. Используется для главного экрана."""
    ensure_state(state)
    ents = [
        e for e in state.get("entries", [])
        if e.get("type") in ("note", "task") and not e.get("done", False)
    ]
    return sorted(ents, key=lambda e: e.get("ts", ""))

def get_all_entries(state: Dict[str, Any]) -> List[Dict[str, Any]]:
    ensure_state(state)
    return sorted(state.get("entries", []), key=lambda e: e.get("ts", ""))

def compute_stats(state: Dict[str, Any]) -> Tuple[int, float, int]:
    ensure_state(state)
    ents = state.get("entries", [])
    tasks = [e for e in ents if e.get("type") == "task" and not e.get("done", False)]
    fins = [e for e in ents if e.get("type") == "finance"]
    notes = [e for e in ents if e.get("type") == "note" and not e.get("done", False)]
    tcnt = len(tasks)
    fsum = sum(float(e.get("amount") or 0) for e in fins)
    ncnt = len(notes)
    return tcnt, fsum, ncnt

def fmt_entry(e: Dict[str, Any]) -> str:
    """Простой форматтер. Можно расширить языком позже."""
    typ = e.get("type", "note")
    text = e.get("text", "")
    assignee = e.get("assignee")
    done = e.get("done", False)

    if typ == "task":
        prefix = "[Задача]"
        if assignee:
            prefix += f" {assignee}"
    elif typ == "finance":
        amt = float(e.get("amount") or 0)
        prefix = f"[Фин {amt:+.2f}]"
    else:
        prefix = "[Заметка]"

    suffix = " ✓" if done else ""
    return f"{prefix} {text}{suffix}"

# ==================== ЦЕНТРАЛИЗОВАННАЯ НУМЕРАЦИЯ ====================
def build_numbered_view(state: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[int, str]]:
    """
    Один источник правды для нумерации.

    Возвращает:
      - список видимых записей в порядке отображения (заметки, потом задачи)
      - карту {номер: id} для команд /del, /done, /e

    Нумерация совпадает с тем, что печатает пользователь (grouped).
    После мутации нужно заново получать view.
    """
    all_vis = get_visible_entries(state)
    notes = [e for e in all_vis if e.get("type") == "note"]
    tasks = [e for e in all_vis if e.get("type") == "task"]
    visible = notes + tasks
    number_map: Dict[int, str] = {}
    for i, e in enumerate(visible, 1):
        number_map[i] = e["id"]
    return visible, number_map

def get_numbered_finances(state: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[int, str]]:
    """Отдельная нумерация для /fin (финансы не в главной нумерации)."""
    fins = [e for e in get_all_entries(state) if e.get("type") == "finance"]
    number_map: Dict[int, str] = {}
    for i, e in enumerate(fins, 1):
        number_map[i] = e["id"]
    return fins, number_map

def get_entry_by_number(state: Dict[str, Any], number_map: Dict[int, str], num: int) -> Dict[str, Any] | None:
    """Безопасное получение записи по номеру из текущей карты."""
    eid = number_map.get(num)
    if not eid:
        return None
    for e in state.get("entries", []):
        if e.get("id") == eid:
            return e
    return None

def parse_bare_input(raw: str) -> Dict[str, Any] | None:
    """
    Smart bare input (по ТЗ):
      - +число текст или -число текст → finance (текст обязателен)
      - @username текст → task
      - остальное → note
    """
    import re
    raw = raw.strip()
    if not raw:
        return None

    # Финансы
    if raw[0] in ('+', '-'):
        m = re.match(r'^([+-]?\s*\d+(?:\.\d+)?)\s+(.+)$', raw)
        if m:
            try:
                amt_str = m.group(1).replace(' ', '')
                amt = float(amt_str)
                text = m.group(2).strip()
                if text:
                    return {"type": "finance", "text": text, "amount": amt}
            except ValueError:
                pass
        return None

    # Задача
    if raw.startswith('@'):
        m = re.match(r'^(@\S+)\s+(.+)$', raw)
        if m:
            assignee = m.group(1)
            text = m.group(2).strip()
            if text:
                return {"type": "task", "text": text, "assignee": assignee}
        return None

    # Заметка
    return {"type": "note", "text": raw}


def get_user_tasks_across_projects(username: str) -> Dict[str, List[Dict[str, Any]]]:
    """Чистая бизнес-логика для бота: задачи по @username по всем проектам."""
    from .data import list_projects, load_state
    result: Dict[str, List[Dict[str, Any]]] = {}
    target = f"@{username}" if not username.startswith('@') else username
    for proj in list_projects():
        try:
            st = load_state(proj["dir"])
            tasks = [
                e for e in st.get("entries", [])
                if e.get("type") == "task"
                and not e.get("done", False)
                and e.get("assignee") == target
            ]
            if tasks:
                result[proj["name"]] = tasks
        except Exception:
            continue
    return result


# ==================== Экспорт (один .xlsx файл с 3 листами) ====================
def export_to_xlsx(state: Dict[str, Any], filepath: Path) -> None:
    """Экспорт в один .xlsx файл с тремя листами: Заметки, Задачи, Финансы.
    Требует openpyxl: pip install openpyxl
    """
    ensure_state(state)
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("Для /export нужен openpyxl. Установи: pip install openpyxl")

    wb = Workbook()

    # Удаляем дефолтный лист
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Заметки
    ws_notes = wb.create_sheet("Заметки")
    ws_notes.append(["Дата", "Текст", "Выполнена"])
    for e in get_all_entries(state):
        if e.get("type") == "note":
            ws_notes.append([
                e.get("ts", ""),
                e.get("text", ""),
                "Да" if e.get("done") else "Нет"
            ])

    # Задачи
    ws_tasks = wb.create_sheet("Задачи")
    ws_tasks.append(["Дата", "Текст", "Ответственный", "Выполнена"])
    for e in get_all_entries(state):
        if e.get("type") == "task":
            ws_tasks.append([
                e.get("ts", ""),
                e.get("text", ""),
                e.get("assignee") or "",
                "Да" if e.get("done") else "Нет"
            ])

    # Финансы
    ws_fins = wb.create_sheet("Финансы")
    ws_fins.append(["Дата", "Сумма", "Пояснение", "Выполнена"])
    for e in get_all_entries(state):
        if e.get("type") == "finance":
            ws_fins.append([
                e.get("ts", ""),
                e.get("amount"),
                e.get("text", ""),
                "Да" if e.get("done") else "Нет"
            ])

    wb.save(filepath)

